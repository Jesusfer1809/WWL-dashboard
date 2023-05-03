from flask import Flask, render_template, request, json, jsonify, session, redirect, send_from_directory

from flask_session import Session

from flask_cors import CORS, cross_origin

from flask_sqlalchemy import SQLAlchemy
# from flask_marshmallow import Marshmallow
from sqlalchemy import text

import kml2geojson
import psycopg2
from datetime import datetime, timedelta, date
#import utilities_functions as uf 
from werkzeug.security import generate_password_hash, check_password_hash
import excel_functions as ef

import os
from werkzeug.utils import secure_filename
from uuid import uuid4
import json

from decimal import Decimal
os.environ['OPENBLAS_NUM_THREADS'] = '1'
from openpyxl import load_workbook


import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from fpdf import FPDF


app = Flask(__name__, static_url_path='', static_folder='wwl-frontend/dist', template_folder='wwl-frontend/dist')
app.secret_key = "SUPERSECRETWWLKEY"


app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://WWL_ADMIN:WWL#2023@wwl-rossing.crnkanilun4m.ap-southeast-2.rds.amazonaws.com/wwlengineering_rossing'
app.config['SECRET_KEY'] = "SUPERSECRETWWLKEY"
app.config['SESSION_SECRET'] = "SUPERSECRETWWLKEY"
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_COOKIE_SAMESITE'] = "None"
app.config['SESSION_COOKIE_SECURE'] = True

app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=20)
app.config['SESSION_COOKIE_NAME'] = 'logged_user'
db = SQLAlchemy(app)

app.secret_key = app.config['SESSION_SECRET']
# ma = Marshmallow(app)


Session(app)
CORS(app)

#app.config.from_pyfile('config.cfg')

BASEPATH = os.getcwd()

dbname = 'wwlengineering_rossing'
user = 'WWL_ADMIN'
password = 'WWL#2023'
host = 'wwl-rossing.crnkanilun4m.ap-southeast-2.rds.amazonaws.com'
dev = True

def dbconnect():
    conn = psycopg2.connect(host=host,
                            database=dbname,
                            user=user,
                            password=password)
    return conn

def get_piezometer_data():
    userQuery = db.session.execute(text(f"SELECT paddock as piezo_paddock ,id as piezo_name, depth as piezo_depth, datalogger as piezo_node, channel as piezo_channel FROM piezometer_details WHERE status = 1;"))
    
    piezo_data = [dict(r._mapping) for r in userQuery]

    # con = dbconnect()
    # cur = con.cursor()
    # query = "select paddock,id as name, depth, datalogger as node, " +\
    #      "channel from piezometer_details where status = 1"
    # cur.execute(query)
    # con.commit()
    # data = cur.fetchall()
    return piezo_data

def query(data_obj,ammount,period):
    # query = "select min(pressure),max(pressure),avg(pressure) "+\
    #         "from node_%s_%s where (current_date - time) <= interval '%s' %s"%(node,channel,ammount,period)
    # con = dbconnect()
    # cur = con.cursor()
    # cur.execute(query)
    # con.commit()
    # vals = cur.fetchall()

    userQuery = db.session.execute(text(f"select min(pressure) as min ,max(pressure) as max ,avg(pressure) as avg from node_{data_obj['piezo_node']}_{data_obj['piezo_channel']} where (current_date - time) <= interval '{ammount}' {period};"))
    piezo_data = [dict(r._mapping) for r in userQuery]
    # print("vals",vals)
    # for x in vals:
    #     data = data + tuple(map(str,list(x)))

    data_dict = piezo_data[0]

    # print("final-query-data",data)
    return [data_dict['min'], data_dict['max'], data_dict['avg']]

def get_data():
    piezo_data = get_piezometer_data()
    new_data = []
    # looping around all piezometers
    for data in piezo_data:
        weekly = query(data,14,'day')
        monthly = query(data,1,'month')
        quarterly = query(data,3,'month')

        lectures_data = [weekly, monthly, quarterly]

        flat_list = list(np.concatenate(lectures_data).flat)

        new_data.append({
           "piezo_depth": data["piezo_depth"],
           "piezo_name": data["piezo_name"],
           "piezo_paddock": data["piezo_paddock"],
           "piezo_lectures":flat_list
        })

    print("new_data", new_data)
    return new_data

@app.route("/api/v1/excel-data", methods=["GET"])
@cross_origin()
def get_excel_data():
    
    # print(BASEPATH)
    # print("UWU")
    
    # tdata = get_data()
    print(os.path.abspath( "pyreport/report.xlsx"))
    data = get_data()



    return jsonify({
        "excel-data":data,
    })

def read_excel():
    
    tdata = get_data()
    
    filename = os.path.abspath( "pyreport/report.xlsx") 
    # filename = BASEPATH + "/pyreport/report.xlsx"
    print(filename)
    
    #print(piezo_data)
    wb = load_workbook(filename)
    print(wb)
    sh = wb.active
    i=14
    j=7
    for data in tdata:
        print(data)
        sh.cell(row=i,column=3).value = data['piezo_paddock']
        sh.cell(row=i,column=5).value = data['piezo_name']
        sh.cell(row=i,column=16).value = data['piezo_depth']

        for val in data['piezo_lectures']:
            sh.cell(row=i,column=j).value = "-" if val is None else float(val)
            j+=1

        # for j in range(3,12):
        #     print(i,j)
        #     sh.cell(row=i,column=4+j).value = float(data[j])
        j=7
        i+=1
    sh.cell(row=5,column=13).value = date.today()

    
    wb.save(os.path.abspath( "../client/public/pyreport/report2.xlsx") )
    return i



class piezometer_details(db.Model):

    __tablename__ = 'piezometer_details'

    id = db.Column(db.String(length=50), primary_key=True) 
    paddock = db.Column(db.String(length=50))
    section = db.Column(db.String(length=50))
    cptu = db.Column(db.String(length=50))

    datalogger = db.Column(db.Integer())
    channel = db.Column(db.Integer())
    serial = db.Column(db.BigInteger())
    depth = db.Column(db.Numeric(10,5))
    lat = db.Column(db.Numeric(20,14))
    lon = db.Column(db.Numeric(20,14))
    status = db.Column(db.Integer(),nullable=False)
   
    def obj_to_dict(self):
        return {"id":self.id,
        "paddock":self.paddock,
        "section":self.section,
        "cptu":self.cptu,
        "datalogger":self.datalogger,
        "channel":self.channel,
        "serial":self.serial,
        "depth":self.depth,
        "lat":self.lat,
        "lon":self.lon,
        "status":self.status
        }
    
    # def __repr__(self):
    #     return f'Item {self.name}'

class Incident_report(db.Model):

    __tablename__ = 'incident_reports'

    id = db.Column(db.String(length=50), primary_key=True)
    from_user = db.Column(db.Integer(), nullable=False)
    photo = db.Column(db.String(length=100), default="incident-default" )
    title= db.Column(db.String(length=50), nullable=False)
    paddock= db.Column(db.String(length=20), nullable=False)
    date= db.Column(db.String(length=20), nullable=False)

    latitude = db.Column(db.Numeric(), nullable=False)
    longitude= db.Column(db.Numeric(), nullable=False)
    elevation= db.Column(db.Numeric(), nullable=False)
    description= db.Column(db.String(length=150), nullable=False)

    supervisors= db.Column(db.ARRAY(db.String(length=40)))
    

    def __init__(self, id, from_user, photo, title, paddock, date, latitude, longitude, elevation, description, supervisors ):
        self.id = id
        self.from_user = from_user
        self.photo = photo
        self.title = title
        self.paddock=paddock
        self.date=date
        self.latitude=latitude
        self.longitude=longitude
        self.elevation=elevation
        self.description=description
        self.supervisors=supervisors
        

    def obj_to_dict(self):
        return {
            "id":self.id,
            "from_user":self.from_user,
            "photo":self.photo,
            "title":self.title,
            "paddock":self.paddock,
            "date":self.date,
            "latitude":self.latitude,
            "longitude":self.longitude,
            "elevation":self.elevation,
            "description":self.description,
            "supervisors":self.supervisors,      
        }


class Piezometer_report(db.Model):

    __tablename__ = 'piezometer_reports'

    id = db.Column(db.String(length=50), primary_key=True)
    from_user = db.Column(db.Integer(), nullable=False)
    photo = db.Column(db.String(length=100), default="piezoreport-default" )
    title= db.Column(db.String(length=50), nullable=False)
    paddock= db.Column(db.String(length=20), nullable=False)
    piezo= db.Column(db.String(length=20), nullable=False)
    date= db.Column(db.String(length=20), nullable=False)
    description= db.Column(db.String(length=150), nullable=False)
    supervisors= db.Column(db.ARRAY(db.String(length=40)))

    def __init__(self, id, from_user,photo, title, paddock, piezo, date, description, supervisors ):
        self.id = id
        self.from_user = from_user
        self.photo = photo
        
        self.title = title
        self.paddock=paddock
        self.piezo = piezo
        self.date = date
        self.description=description
        self.supervisors=supervisors

    def obj_to_dict(self):
        return {
            "id":self.id,
            "from_user":self.from_user,
            "photo":self.photo,
            "title":self.title,
            "paddock":self.paddock,
            "piezo":self.piezo,
            "date":self.date,
            "description":self.description,
            "supervisors":self.supervisors,
        }
        

class Last_readings(db.Model):

    __tablename__ = 'last_readings'

    node = db.Column(db.Integer(), primary_key=True) 
    channel = db.Column(db.Integer(), primary_key=True)
    time = db.Column(db.DateTime())

    pressure = db.Column(db.Numeric(12,6))
    temperature = db.Column(db.Numeric(12,6))
    
    def obj_to_dict(self):
        return {"node":self.node,
        "channel":self.channel,
        "time":self.time,
        "pressure":self.pressure,
        "temperature":self.temperature,
        }
    
    # def __repr__(self):
    #     return f'Item {self.name}'

class PDF(FPDF):
    def header(self):
        self.image(os.path.abspath("../client/public/media/img/photos") + "\\" + "rossing_logo.png", 10,8, 30)
        
        self.image(os.path.abspath("../client/public/media/img/photos") + "\\" + "wwl-black.png", (self.w -40) ,14, 30)

        self.ln(40)

last_node = 0

def dict_helper(objlist):
    result2 = [item.obj_to_dict() for item in objlist]
    return result2



@app.route('/api/v1/piezometers-data', methods=['GET'])
@cross_origin()
def getUsers():
    users=[]
    
    result = piezometer_details.query.all()
    
    
    piezos = dict_helper(result)
    

    
    return jsonify({
        "message":"success",
        "results": len(piezos),
        "piezos": piezos
    })

@app.route('/api/v1/last-readings', methods=['GET'])
@cross_origin()
def get_last_readings():
    users=[]
    
    result = Last_readings.query.all()
    
    readings = dict_helper(result)
    

    
    return jsonify({
        "message":"success",
        "results": len(readings),
        "readings": readings
    })

@app.route('/api/v1/piezometers-data/<paddock>/<piezo>', methods=['GET'])
@cross_origin()
def getOnePiezoInfo(paddock,piezo):
    users=[]
    fixed_paddock = paddock

    if paddock == 'E1-E2':
        fixed_paddock = 'E1/E2'
    elif paddock == 'Y1-Y2':
        fixed_paddock = 'Y1/Y2'
    else:
        fixed_paddock = paddock
    
    
    result = piezometer_details.query.filter_by(paddock=fixed_paddock,id=piezo).all()
    
    
    piezos = dict_helper(result)
    

    
    return jsonify({
        "message":"success",
        "piezos":piezos       
    })



# GET THE LECTURES OF A GIVEN PIEZOMETER IN AN INTERVAL OF DAYS
@app.route('/api/v1/lectures/<node>/<daysAgo>', methods=['GET'])
@cross_origin()
def getLectures(node,daysAgo):
    d = datetime.today() - timedelta(days=int(daysAgo))
    date = d.strftime("%Y-%m-%d 00:00:00")
    # connection = db.session.connection()
    result = db.session.execute(text(f"SELECT time,pressure FROM public.{node} WHERE time >= '{date}'"))
    

    lectures = [dict(r._mapping) for r in result]

    return jsonify({"message":"success","lectures":lectures})




@app.route("/api/v1/login", methods=["POST"])
@cross_origin()
def login_user():
    attempted_username = request.json["username"]
    attempted_password = request.json["password"]

    userQuery = db.session.execute(text(f"SELECT * FROM public.users WHERE username = '{attempted_username}'"))
    
    result = [dict(r._mapping) for r in userQuery]

    print('results',result)

    if len(result) == 0:
        return jsonify({"error":"Unauthorized"}),401

    user = result[0]

    print("USER",user)

    if not check_password_hash(user['password'],attempted_password):
        return jsonify({"error":"Unauthorized"}),401

    session["user_id"] = user['user_id']

    return jsonify({"message":"logged in!", "name":user["name"]}) 


@app.route("/api/v1/logout", methods=["POST"])
@cross_origin()
def logout():
    session.pop("user_id")
    return jsonify({"message":"logged out!"})

@app.route("/api/v1/current-user", methods=["GET"])
@cross_origin()
def current_user_info():
    user_id = session.get('user_id')
    print('SESSION',user_id)

    if not user_id:
        return jsonify({"error":"Unauthorized"}), 401

    userQuery = db.session.execute(text(f"SELECT * FROM public.users WHERE user_id = '{user_id}'"))
    
    result = [dict(r._mapping) for r in userQuery]

    user = result[0]

    return jsonify({
        "id": user["user_id"],
        "username": user["username"],
        "name": user["name"],
        "picture": user["picture"],
        "user_id": user_id
    })

@app.route("/api/v1/user/<id>", methods=["GET"])
@cross_origin()
def get_user_from_id(id):


    userQuery = db.session.execute(text(f"SELECT * FROM public.users WHERE user_id = '{id}'"))
    
    result = [dict(r._mapping) for r in userQuery]

    user = result[0]

    return jsonify({
        "id": user["user_id"],
        "username": user["username"],
        "name": user["name"],
        "picture": user["picture"],
        "user_id": id
    })


def upload_photo():
 
    if len(request.files.to_dict(flat=False)) == 0:
        return "incident-default"
    
    file = request.files['photo']
    
    filename = secure_filename(file.filename)

    final_filename = f'{uuid4()}-{filename}'

    file.save(os.path.abspath( "../client/public/media/incident_reports") + "\\" + final_filename)
    return final_filename

def upload_piezoreport_photo():

    if len(request.files.to_dict(flat=False)) == 0:
        return "piezoreport-default"
    
    file = request.files['photo']
    # print(file)
    # print("name",file.filename)

    filename = secure_filename(file.filename)

    final_filename = f'{uuid4()}-{filename}'

    # print(os.path.abspath( "../client/public/media/piezometer_reports") + "\\" + final_filename)


    file.save(os.path.abspath( "../client/public/media/piezometer_reports") + "\\" + final_filename)
    return final_filename

    


@app.route("/api/v1/incident-reports", methods=["GET"])
@cross_origin()
def get_incidents():
    
    userQuery = db.session.execute(text(f"SELECT ir.id as incident_id, ir.title as incident_title, ir.photo as incident_photo,  ir.paddock as incident_paddock, ir.date as incident_date, ir.latitude as incident_latitude, ir.longitude as incident_longitude, ir.elevation as incident_elevation, ir.description as incident_description, ir.supervisors as incident_supervisors, u.username as user_username , u.user_id, u.name as user_name, u.picture user_picture FROM incident_reports as ir LEFT JOIN users as u ON ir.from_user = u.user_id;"))
    
    incident_reports = [dict(r._mapping) for r in userQuery]


    return jsonify({
        "message":"success",
        "results": len(incident_reports),
        "incidents": incident_reports
    })

@app.route("/api/v1/piezometer-reports", methods=["GET"])
@cross_origin()
def get_piezo_reports():


    userQuery = db.session.execute(text(f"SELECT pr.id as report_id, pr.title as report_title, pr.photo as report_photo, pr.paddock as report_paddock, pr.piezo as report_piezo, pr.date as report_date, pr.description as report_description, pr.supervisors as report_supervisors,  u.username as user_username , u.user_id, u.name as user_name, u.picture user_picture FROM piezometer_reports as pr LEFT JOIN users as u ON pr.from_user = u.user_id;"))
    
    piezo_reports = [dict(r._mapping) for r in userQuery]


    return jsonify({
        "message":"success",
        "results": len(piezo_reports),
        "reports": piezo_reports
    })

@app.route('/api/v1/new-incident-report', methods=['POST'])
@cross_origin()
def new_incident_report():
    
    
    from_user = request.form["from_user"]
    
    title = request.form["title"]
    paddock= request.form["paddock"]
    date= request.form["date"]
    latitude= request.form["latitude"]
    longitude= request.form["longitude"]
    elevation= request.form["elevation"]
    description=request.form["description"]
    supervisors=request.form["supervisors"].split(",")


    photo_db = upload_photo()
    report_id = uuid4()

    new_report = Incident_report(report_id, from_user, photo_db, title, paddock, date, latitude, longitude, elevation, description, supervisors )

    db.session.add( new_report )
    db.session.commit()
    return jsonify({"status":"success"})


@app.route('/api/v1/new-piezometer-report', methods=['POST'])
@cross_origin()
def new_piezometer_report():
    
    from_user = request.form["from_user"]
    
    title = request.form["title"]
    
    paddock= request.form["paddock"]
    piezo= request.form["piezo"]
    date= request.form["date"]
    description=request.form["description"]
    
    supervisors=request.form["supervisors"].split(",")
    print("supervisors",supervisors)
    
    
    photo_db = upload_piezoreport_photo()
    
    report_id = uuid4()
    
    new_report = Piezometer_report(report_id, from_user, photo_db, title, paddock, piezo, date, description, supervisors)

    db.session.add( new_report )
    db.session.commit()
    return jsonify({"status":"success"})


@app.route('/api/v1/incident-reports/<id>', methods=['GET'])
@cross_origin()
def getOneIncident(id):
    
    userQuery = db.session.execute(text(f"SELECT ir.id as incident_id, ir.title as incident_title, ir.photo as incident_photo,  ir.paddock as incident_paddock, ir.date as incident_date, ir.latitude as incident_latitude, ir.longitude as incident_longitude, ir.elevation as incident_elevation, ir.description as incident_description, ir.supervisors as incident_supervisors, u.username as user_username , u.user_id, u.name as user_name, u.picture user_picture FROM incident_reports as ir LEFT JOIN users as u ON ir.from_user = u.user_id WHERE ir.id = '{id}';"))
    
    incident_reports = [dict(r._mapping) for r in userQuery]


    if(len(incident_reports) == 0):
        return jsonify({
            "message": "error"
        }), 404


    return jsonify({
        "message":"success",
        "report": incident_reports[0]
    })

@app.route('/api/v1/piezometer-reports/<id>', methods=['GET'])
@cross_origin()
def getOnePiezoReport(id):

    userQuery = db.session.execute(text(f"SELECT pr.id as report_id, pr.title as report_title, pr.photo as report_photo, pr.paddock as report_paddock, pr.piezo as report_piezo, pr.date as report_date, pr.description as report_description, pr.supervisors as report_supervisors,  u.username as user_username , u.user_id, u.name as user_name, u.picture user_picture FROM piezometer_reports as pr LEFT JOIN users as u ON pr.from_user = u.user_id WHERE pr.id = '{id}';"))
    
    piezo_reports = [dict(r._mapping) for r in userQuery]

    if(len(piezo_reports) == 0):
        return jsonify({
            "message": "error"
        }), 404


    return jsonify({
        "message":"success",
        "report": piezo_reports[0]
    })


@app.route('/api/v1/get_geojson_<folder>-<name>', methods=['GET'])
def get_geojson(folder,name):
    
        name = 'data/'+folder+'/'+name.upper()+'.kml'
        data = kml2geojson.main.convert(name)
        return jsonify(data)
    # else:
    #     return render_template('no_authorized.html')


@app.route('/api/v1/modify_excel', methods=['POST'])
def modify_excel():
    
    now = datetime.now()
    dt_string = now.strftime("%Y/%m/%d %H:%M:%S")
    data = read_excel()
    print("report download by %s at %s"%(session.get('user_id'),dt_string))
    return jsonify({
        "filename":os.path.abspath( "pyreport/report2.xlsx")
    })



def create_chart(paddock, piezo, days, initial_pressure,dates):
    def get_pressure_as_float(p):
        return float(p)

    # def getDate(obj): 
    #     return obj['time']

    pressure = list(map(get_pressure_as_float, initial_pressure))
    # dates = list(map(getDate, lectures))

    # Data for plotting
    t = dates
    s = pressure

    plt.plot(t,s)

    plt.xlabel('Dates')
    plt.ylabel('Pressure (KPa)')
    plt.xticks(visible=False)
    plt.fill_between(t, s, min(pressure), color=['#477C9A'], alpha=.1)

    plt.grid()

    now = datetime.now()
    dt_string = now.strftime("%Y%m%d%H%M%S")

    
    filename = os.path.abspath( "../client/public/media/charts") + "\\" + f"{paddock}_{piezo}_{days}_{dt_string}.png"
    chart_filename = f"{paddock}_{piezo}_{days}_{dt_string}.png"
    print("FILENAME", filename)

    
    plt.savefig(filename)
    return chart_filename



def create_pdf(title,description,paddock, piezo, date, averagePWP, inoperativeDates,days, chart_filename, sectionURL, lecturesDates):
    pdf = PDF('P','mm','Letter')

    pdf.set_auto_page_break(auto=True, margin=15)
    #Add a page
    pdf.add_page()
    pdf.set_text_color(34,34,34)


    #TITLE
    pdf.set_font('helvetica','',20)
    pdf.cell(0,10,title)
    pdf.ln(18)

    #SEPARATION LINE
    pdf.set_fill_color(241,245,249)
    pdf.cell(0,0.5,'', fill=True)
    pdf.ln(18)


    #DESCRIPTION
    pdf.set_font('helvetica','',12)
    pdf.cell(0,10,description)
    pdf.ln(24)

    #SUBTITLE
    pdf.set_font('helvetica','U',16)
    pdf.cell(0,10,'Piezo information')
    pdf.ln(18)


    pdf.set_font('helvetica','',14)
    pdf.cell(50,10,'Paddock section: ')

    pdf.set_font('helvetica','',12)
    pdf.cell(20,10,paddock)
    pdf.ln(20)


    pdf.set_font('helvetica','',14)
    pdf.cell(50,10,'Piezometer ID: ')

    pdf.set_font('helvetica','',12)
    pdf.cell(20,10,piezo)
    pdf.ln(20)


    pdf.set_font('helvetica','',14)
    pdf.cell(50,10,'Inspection date: ')

    pdf.set_font('helvetica','',12)
    pdf.cell(20,10,date)
    pdf.ln(20)


    if len(lecturesDates) !=0:
        pdf.set_font('helvetica','',14)
        pdf.cell(50,10, f'Average PWP (From {lecturesDates[0]} to {lecturesDates[len(lecturesDates)-1]}): ')

        pdf.ln(12)

        pdf.set_font('helvetica','',12)
        pdf.cell(20,10, f'{averagePWP} KPa')
        pdf.ln(20)
    else:
        pdf.set_font('helvetica','',14)
        pdf.cell(20,10, f'No recent lectures for {days} days span')
        pdf.ln(20)

    if len(inoperativeDates) != 0:
        pdf.set_font('helvetica','',14)
        pdf.cell(50,10, f'Inoperative dates (From {inoperativeDates[0]["currentDate"]} to today): ')

        pdf.ln(12)

        pdf.set_font('helvetica','',12)

        for obj in inoperativeDates:        
            pdf.cell(20,10,f'{obj["inoperativeDays"]} days: from {obj["currentDate"]} to {obj["nextDate"]}')
        

    pdf.ln(500)

    if chart_filename != "None":
        #SUBTITLE
        pdf.set_font('helvetica','U',16)
        pdf.cell(0,10,f'Latest piezometer lectures ( last {days} days )')
        pdf.ln(20)

        pdf.image(os.path.abspath("../client/public/media/charts") + "\\" + chart_filename,10,60,pdf.w-20)
        pdf.ln(500)

    if sectionURL != "None":
    #SUBTITLE
        pdf.set_font('helvetica','U',16)
        pdf.cell(0,10,'Section graph')
        pdf.ln(20)

    #SECTION IMG
    

        pdf.image(os.path.abspath("../client/public/media/img/sections") + "\\" + sectionURL,10,80,pdf.w-20)

    now = datetime.now()
    dt_string = now.strftime("%Y%m%d%H%M%S")

    
    save_filename = os.path.abspath("../client/public/report_pdf") + "\\" + f"{paddock}_{piezo}_{days}_{dt_string}.pdf"
    download_filename = f"{paddock}_{piezo}_{days}_{dt_string}.pdf"
    pdf.output(save_filename)

    return download_filename


@app.route('/api/v1/create-pdf', methods=['POST'])
def save_pdf():

    title = request.json["title"]
    description = request.json["description"]
    paddock = request.json["paddock"]
    piezo = request.json["piezo"]
    days = request.json["days"]
    date = request.json["date"]
    averagePWP = request.json["averagePWP"]
    inoperativeDates = request.json["inoperativeDates"]
    lecturesDates = request.json["lecturesDates"]
    lecturesPressure = request.json["lecturesPressure"]
    sectionURL = request.json["sectionURL"]
    photo = request.json["photo"]
    supervisors=request.json["supervisors"].split(",")

    # print({"supervisors":supervisors,
    #        "title":title,
    #        "description":description,
    #        "paddock":paddock,
    #        "piezo":piezo,
    #        "days":days,
    #        "date":date,
    #        "averagePWP":averagePWP,
    #        "inoperativeDates":inoperativeDates,
    #        "lecturesPressure":lecturesPressure,
    #        "sectionURL":sectionURL,
    #        "photo":photo
    #        })

    # print("lecturesDates", lecturesDates)
  
    if len(lecturesDates) != 0 and len(lecturesPressure) != 0:
        chart_filename = create_chart(paddock, piezo, days,lecturesPressure,lecturesDates)
        filename = create_pdf(title, description, paddock, piezo, date, averagePWP, inoperativeDates, days, chart_filename, sectionURL, lecturesDates)
        
        return jsonify({
            "filename": filename,
            # "chart_filename": chart_filename
        })
    else:
        chart_filename ="None"
        filename = create_pdf(title, description, paddock, piezo, date, averagePWP, inoperativeDates, days, chart_filename, sectionURL, lecturesDates)
        return jsonify({
            "filename": filename,
        })
    
    

   

# @app.route("/")
# @cross_origin()
# def main():    
#     "MAIN"
#     return render_template("index.html")



# send_from_directory(app.static_folder,'wwl-frontend/dist/index.html')
    